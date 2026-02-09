#!/usr/bin/env python3


import sqlite3
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

"""
Collection and validation of information like assignment name, cat(category
weight, and grades
"""
#arr = []   # Stores all assignments as dictionaries

def create_archive_folder():
    """Create archives folder if it doesn't exist"""
    if not os.path.exists("archives"):
        os.makedirs("archives")
        print("✓ Created 'archives' folder")


def setup_db():
    """Create database connection and table"""

    #To add timestamp to the database name as a way to filter different databases
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    db_name = f"archives/grades_{timestamp}.db"


    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assignment_name TEXT NOT NULL,
            grade REAL NOT NULL,
            category TEXT NOT NULL CHECK(category IN ('FA', 'SA')),
            weight REAL NOT NULL,
            weighted_grade REAL NOT NULL
        )
    ''')
    conn.commit()
    return conn, cursor, db_name

def archive_db(db_name):
    """Move databases to archives folder"""
    import shutil #shutil is a module that allows you to move files and directories.
    
    #Make sure the archives folder exists
    create_archive_folder()
    # Move the database file to the archives folder
    source = db_name
    destination = os.path.join("archives", os.path.basename(db_name))
    shutil.move(source, destination)


def info(conn, cursor):
    """Collect and validate assignment info"""
    
    # Assignment name input

    Aname = input("Enter assignment name: ")

    # Category validation
    while True:
        cat = input("Enter category (FA for Formative, SA for Summative): ").upper()
        if cat not in ["FA", "SA"]:
            print("Invalid category! Please type FA or SA.")
        else:
            break

    # Grade validation
    while True:
        try:
            grade = float(input("Enter grade (0 - 100): "))
            if 0 <= grade <= 100:
                break
            else:
                print("Grade must be between 0 and 100.")
        except ValueError:
            print("Invalid input. Enter a number.")

    # Weight validation
    while True:
        try:
            weight = float(input("Enter assignment weight: "))
            if weight > 0:
                break
            else:
                print("Weight must be a positive number.")
        except ValueError:
            print("Invalid input. Enter a number.")
        #calcutation and display of grades out of the given weight.
    weighted_grade = (grade / 100) * weight
    print(f"Weighted Grade: {weighted_grade}")


# Initialize database (How the table is created)

    
    cursor.execute('''INSERT INTO grades (assignment_name, category, grade, weight, weighted_grade) VALUES (?, ?, ?, ?, ?)''',
               (Aname, cat, grade, weight, weighted_grade))
    conn.commit()
    print("✓ Assignment added successfully!\n")

create_archive_folder()

conn, cursor, db_name = setup_db()

def export_to_excel(cursor, total_FA, total_weight_FA, total_SA, total_weight_SA, 
                    final_grade, gpa, promoted, timestamp):
    """Export all data to a formatted Excel file"""
    
    # Create a new workbook
    wb = Workbook()
    
    # ========== SHEET 1: ASSIGNMENTS ==========
    ws1 = wb.active
    ws1.title = "Assignments"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Create headers
    headers = ["Assignment Name", "Category", "Grade", "Weight", "Weighted Grade"]
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Get all assignments from database
    cursor.execute("SELECT assignment_name, category, grade, weight, weighted_grade FROM grades")
    assignments = cursor.fetchall()
    
    # Add data rows
    for row_num, assignment in enumerate(assignments, 2):
        for col_num, value in enumerate(assignment, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = center_align
            
            # Color code based on grade (column 3)
            if col_num == 3:  # Grade column
                if value < 50:
                    for c in range(1, 6):  # Color entire row
                        ws1.cell(row=row_num, column=c).fill = fail_fill
                else:
                    for c in range(1, 6):
                        ws1.cell(row=row_num, column=c).fill = pass_fill
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 15
    
    # ========== SHEET 2: SUMMARY ==========
    ws2 = wb.create_sheet(title="Summary")
    
    # Title
    ws2['A1'] = "GRADE SUMMARY REPORT"
    ws2['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws2.merge_cells('A1:B1')
    
    # FA Section
    ws2['A3'] = "Formative Assessments (FA):"
    ws2['A3'].font = Font(bold=True, size=12)
    ws2['A4'] = "Total Points Earned:"
    ws2['B4'] = f"{total_FA:.2f}"
    ws2['A5'] = "Total Possible Points:"
    ws2['B5'] = f"{total_weight_FA:.2f}"
    
    # SA Section
    ws2['A7'] = "Summative Assessments (SA):"
    ws2['A7'].font = Font(bold=True, size=12)
    ws2['A8'] = "Total Points Earned:"
    ws2['B8'] = f"{total_SA:.2f}"
    ws2['A9'] = "Total Possible Points:"
    ws2['B9'] = f"{total_weight_SA:.2f}"
    
    # Final Results
    ws2['A11'] = "FINAL RESULTS"
    ws2['A11'].font = Font(bold=True, size=14, color="4472C4")
    ws2['A12'] = "Final Grade:"
    ws2['B12'] = f"{final_grade:.2f}%"
    ws2['B12'].font = Font(bold=True, size=12)
    ws2['A13'] = "GPA:"
    ws2['B13'] = f"{gpa:.2f} / 5.0"
    ws2['B13'].font = Font(bold=True, size=12)
    ws2['A14'] = "Status:"
    ws2['B14'] = promoted
    
    # Color code the status
    if promoted == "Pass":
        ws2['B14'].fill = pass_fill
        ws2['B14'].font = Font(bold=True, color="006100")
    else:
        ws2['B14'].fill = fail_fill
        ws2['B14'].font = Font(bold=True, color="9C0006")
    
    # Assignments to resubmit
    cursor.execute("SELECT assignment_name, grade FROM grades WHERE grade < 50")
    redo = cursor.fetchall()
    
    ws2['A16'] = "Assignments to Resubmit:"
    ws2['A16'].font = Font(bold=True, size=12, color="9C0006")
    
    if redo:
        row = 17
        for assignment in redo:
            ws2[f'A{row}'] = f"• {assignment[0]}"
            ws2[f'B{row}'] = f"Grade: {assignment[1]}"
            ws2[f'A{row}'].fill = fail_fill
            ws2[f'B{row}'].fill = fail_fill
            row += 1
    else:
        ws2['A17'] = "✓ No resubmissions needed!"
        ws2['A17'].fill = pass_fill
        ws2['A17'].font = Font(color="006100")
    
    # Adjust column widths
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    
    # Save the file
    excel_filename = f"grades_report_{timestamp}.xlsx"
    wb.save(excel_filename)
    print(f"📊Excel report saved to '{excel_filename}'")
    
    return excel_filename

# MAIN LOOP
while True:
    info(conn, cursor)
    ans = input("Add another assignment? (yes/no): ").lower()
    if ans != "yes":
        break


#calculations for both formative and summative assessments
cursor.execute("SELECT SUM(weighted_grade), SUM(weight) FROM grades WHERE category = 'FA'")
fa_reslt = cursor.fetchone()

total_FA = fa_reslt[0] if fa_reslt[0] is not None else 0
#it means if there are no formative assessments, set total to 0 bit the [0] is index for the first column
total_weight_FA = fa_reslt[1] if fa_reslt[1] is not None else 0

cursor.execute("SELECT SUM(weighted_grade), SUM(weight) FROM grades WHERE category = 'SA'")
sa_reslt = cursor.fetchone()

total_SA = sa_reslt[0] if sa_reslt[0] is not None else 0
total_weight_SA = sa_reslt[1] if sa_reslt[1] is not None else 0

#main calculations both sa and fa to find gpa
total_grade = total_FA + total_SA
total_weight = total_weight_FA + total_weight_SA

if total_weight > 0:
    final_grade = (total_grade / total_weight) * 100
    gpa = (final_grade / 100) * 5
else:
    final_grade = 0
    gpa = 0

# Determine pass/fail by checking of the students has atleast half the >
if total_FA >= (total_weight_FA / 2) and total_SA >= (total_weight_SA / 2):
    promoted = "Pass"
else:
    promoted = "Fail"


# ------------ OUTPUT RESULTS -------------
print("--------- RESULTS ------------")
print(f"Total Formative: {total_FA:.2f} / {total_weight_FA}")
print(f"Total Summative: {total_SA:.2f} / {total_weight_SA}")
print("------------------------------")
print(f"Final Grade: {final_grade:.2f} / 100")
print(f"GPA: {gpa:.2f} / 5")
print(f"Status: {promoted}")


# Identify assignments to redo
cursor.execute("SELECT assignment_name, grade FROM grades WHERE grade < 50")
redo = cursor.fetchall()
if redo:
    print("Assignments to resubmit:")
    for r in redo:
        print(f"Resubmission: {r['assignment_name']} ")
else:
    print("No resubmission needed.")
                                                        
#--save to csv

# Get all data from database
cursor.execute("SELECT assignment_name, category, grade, weight FROM grades")
all_grades = cursor.fetchall()

# Create CSV filename with timestamp
csv_filename = f"grades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

with open(csv_filename, "w", newline="") as file:
    writer = csv.writer(file)
    writer.writerow(["Assignment", "Category", "Grade", "Weight"])
    
    for a in all_grades:
        writer.writerow([a[0], a[1], a[2], a[3]])

print(f"\n Results saved to '{csv_filename}'")

# Close database
conn.close()
# Archive the database
archive_db(db_name)